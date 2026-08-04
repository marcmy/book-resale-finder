from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import SearchResult

_HEADERS = ["ASIN", "Title", "Best Price", "Condition", "Listing URL"]
_REQUESTED_WIDTH_INCHES = {
    "A": 1.5,
    "B": 5.0,
    "C": 1.5,
    "D": 1.5,
    "E": 8.0,
}


def inches_to_excel_width(inches: float, dpi: int = 96) -> float:
    """Approximate an inch width using Excel's character-based column units."""
    pixels = inches * dpi
    if pixels <= 12:
        return pixels / 12
    return (pixels - 5) / 7


def write_results_xlsx(results: list[SearchResult], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    sheet.append(_HEADERS)
    for result in results:
        sheet.append(result.to_output_row())

    header_fill = PatternFill("solid", fgColor="167447")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D8DEE8")
    body_border = Border(bottom=thin)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 24

    for row in sheet.iter_rows(min_row=2, max_col=5):
        for cell in row:
            cell.border = body_border
            cell.alignment = Alignment(vertical="top")
        row[0].number_format = "@"
        row[1].alignment = Alignment(vertical="top", wrap_text=True)
        row[2].number_format = '$0.00'
        row[2].alignment = Alignment(horizontal="right", vertical="top")
        row[3].alignment = Alignment(vertical="top", wrap_text=True)
        row[4].alignment = Alignment(vertical="top", wrap_text=True)
        if row[4].value:
            row[4].hyperlink = str(row[4].value)
            row[4].style = "Hyperlink"

    for column, inches in _REQUESTED_WIDTH_INCHES.items():
        sheet.column_dimensions[column].width = inches_to_excel_width(inches)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:E{max(sheet.max_row, 1)}"
    sheet.sheet_view.zoomScale = 90
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

    if sheet.max_row >= 2:
        table = Table(displayName="BookResaleResults", ref=f"A1:E{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    workbook.save(output_path)
