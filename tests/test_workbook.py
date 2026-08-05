import csv
from pathlib import Path

from openpyxl import load_workbook
import pytest

from book_resale_finder.models import SearchResult
from book_resale_finder.workbook import (
    ebay_search_url,
    inches_to_excel_width,
    write_results_csv,
    write_results_xlsx,
)


def test_csv_output_is_plain_values(tmp_path: Path):
    output = tmp_path / "results.csv"
    write_results_csv(
        [
            SearchResult(
                asin="9780306406157",
                title="A test book",
                best_price=12.34,
                condition="Very Good",
                listing_url="https://www.ebay.com/itm/123",
            )
        ],
        output,
    )
    with output.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    assert rows[0] == ["ASIN", "Title", "Best Price", "Condition", "Listing URL"]
    assert rows[1] == [
        "9780306406157",
        "A test book",
        "12.34",
        "Very Good",
        "https://www.ebay.com/itm/123",
    ]


def test_unavailable_results_keep_asin_and_leave_result_cells_blank(tmp_path: Path):
    results = [
        SearchResult(
            asin="9780000000001",
            title="No matching listing found",
            condition="No match",
            note="No active eBay listing matched this identifier.",
        ),
        SearchResult(
            asin="9780000000002",
            title="Lookup failed",
            condition="Error",
            note="Temporary API failure",
        ),
        SearchResult(
            asin="9780000000003",
            title="Not processed",
            condition="Skipped",
            note="Quota reserve reached",
        ),
    ]

    csv_output = tmp_path / "results.csv"
    write_results_csv(results, csv_output)
    with csv_output.open("r", encoding="utf-8-sig", newline="") as handle:
        csv_rows = list(csv.reader(handle))
    assert csv_rows[1:] == [
        ["9780000000001", "", "", "", ""],
        ["9780000000002", "", "", "", ""],
        ["9780000000003", "", "", "", ""],
    ]

    xlsx_output = tmp_path / "results.xlsx"
    write_results_xlsx(results, xlsx_output)
    workbook = load_workbook(xlsx_output)
    sheet = workbook["Results"]
    assert [sheet.cell(2, column).value for column in range(1, 6)] == [
        "9780000000001",
        None,
        None,
        None,
        None,
    ]
    assert [sheet.cell(3, column).value for column in range(1, 6)] == [
        "9780000000002",
        None,
        None,
        None,
        None,
    ]
    assert [sheet.cell(4, column).value for column in range(1, 6)] == [
        "9780000000003",
        None,
        None,
        None,
        None,
    ]


def test_workbook_headers_widths_and_links(tmp_path: Path):
    output = tmp_path / "results.xlsx"
    write_results_xlsx(
        [
            SearchResult(
                asin="9780306406157",
                title="A test book",
                best_price=12.34,
                condition="Very Good",
                listing_url="https://www.ebay.com/itm/123",
            )
        ],
        output,
    )
    workbook = load_workbook(output)
    sheet = workbook["Results"]
    assert [sheet.cell(1, col).value for col in range(1, 6)] == [
        "ASIN",
        "Title",
        "Best Price",
        "Condition",
        "Listing URL",
    ]
    assert sheet.column_dimensions["A"].width == pytest.approx(inches_to_excel_width(1.5))
    assert sheet.column_dimensions["B"].width == pytest.approx(inches_to_excel_width(5.0))
    assert sheet.column_dimensions["E"].width == pytest.approx(inches_to_excel_width(8.0))
    assert sheet["A2"].hyperlink.target == ebay_search_url("9780306406157")
    assert sheet["C2"].number_format == "$0.00"
    assert sheet["E2"].hyperlink.target == "https://www.ebay.com/itm/123"
    assert sheet.freeze_panes == "A2"
