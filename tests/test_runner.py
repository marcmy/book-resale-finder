import asyncio
from pathlib import Path

from openpyxl import load_workbook

from book_resale_finder.models import QuotaInfo, SearchResult
from book_resale_finder.runner import run_scan


class FakeClient:
    calls = []

    def __init__(self, client_id, client_secret, config):
        self.api_calls = 0

    async def __aenter__(self):
        return self

    async def __aexit__(self, exc_type, exc, tb):
        return None

    async def find_best_listing(self, identifier, *, include_shipping):
        self.api_calls += 1
        self.calls.append(identifier.primary_value)
        return SearchResult(
            asin=identifier.original,
            title=f"Title {identifier.primary_value}",
            best_price=10.0,
            condition="Good",
            listing_url="https://www.ebay.com/itm/1",
        )

    async def fetch_quota(self):
        return QuotaInfo(limit=5000, used=2, remaining=4998, reset_at="2026-08-05T07:00:00Z")


def test_runner_deduplicates_calls_but_preserves_rows(tmp_path: Path, monkeypatch):
    input_file = tmp_path / "masterlist.csv"
    input_file.write_text(
        "ASIN\n9780306406157\n9780306406157\nB000HCWBCG\n",
        encoding="utf-8",
    )
    FakeClient.calls = []
    monkeypatch.setattr("book_resale_finder.runner.EbayClient", FakeClient)

    progress = []
    summary = asyncio.run(
        run_scan(
            input_file=input_file,
            output_dir=tmp_path / "output",
            config={"output_prefix": "results", "max_workers": 2},
            client_id="id",
            client_secret="secret",
            include_shipping=False,
            progress_callback=progress.append,
        )
    )

    assert len(FakeClient.calls) == 2
    assert summary.total_identifiers == 3
    assert summary.unique_identifiers == 2
    assert summary.found == 3
    assert progress[-1].completed == 3
    assert progress[-1].total == 3

    workbook = load_workbook(summary.output_file)
    sheet = workbook["Results"]
    assert sheet.max_row == 4
    assert [sheet.cell(row, 1).value for row in range(2, 5)] == [
        "9780306406157",
        "9780306406157",
        "B000HCWBCG",
    ]
